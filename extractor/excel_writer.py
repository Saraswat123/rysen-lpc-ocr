"""
Writes extracted LPC records to Excel.
One workbook, 4 sheets (one per stage).
Each row = one student × one month × one domain (flat, pivot-friendly).
"""
import pathlib
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.lpc import LPCRecord, DOMAINS_BY_STAGE

SHEET_NAMES = {
    "early_years":           "Early Years (Nursery-KG)",
    "foundational_primary":  "Foundational Primary (Gr 1-2)",
    "preparatory":           "Preparatory Stage (Gr 3-5)",
    "middle":                "Middle Stage (Gr 6-8)",
}

COLUMNS = [
    "Student Name",
    "Class & Sec",
    "Roll No.",
    "Month",
    "Domain",
    "C1",
    "C2",
    "C3",
    "C4",
    "Observational Anecdote",
    "Strengths",
    "Focus for Next Month",
    "Parent Sign",
    "Teacher Sign",
    "Principal Sign",
    "Source PDF",
]

# RYSEN brand colours
HDR_FILL   = PatternFill("solid", fgColor="1D4E6B")   # dark teal
ALT_FILL   = PatternFill("solid", fgColor="EEF4F7")   # light blue-grey
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
DATA_FONT  = Font(size=9)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN       = Side(style="thin", color="CCCCCC")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SCORE_COLS = {"C1", "C2", "C3", "C4"}
WIDE_COLS  = {"Observational Anecdote", "Strengths", "Focus for Next Month"}

SCORE_COLOR = {
    3: "C6EFCE",  # green
    2: "FFEB9C",  # yellow
    1: "FFC7CE",  # red-light
    0: "D3D3D3",  # grey
}


def _header_row(ws, columns: list[str]):
    ws.append(columns)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = HDR_FILL
        cell.font   = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 28


def _col_widths(ws, columns: list[str]):
    for col_idx, col_name in enumerate(columns, 1):
        letter = get_column_letter(col_idx)
        if col_name in WIDE_COLS:
            ws.column_dimensions[letter].width = 30
        elif col_name in SCORE_COLS:
            ws.column_dimensions[letter].width = 6
        elif col_name in {"Student Name", "Class & Sec"}:
            ws.column_dimensions[letter].width = 18
        elif col_name == "Source PDF":
            ws.column_dimensions[letter].width = 28
        else:
            ws.column_dimensions[letter].width = 14
    ws.freeze_panes = "A2"


def _write_record(ws, record: LPCRecord, row_start: int) -> int:
    """Write all rows for one student. Returns next available row."""
    current_row = row_start
    alt = False

    sig = lambda b: "✓" if b else ("✗" if b is False else "")

    for month_entry in record.months:
        month = month_entry.month
        for domain, d in month_entry.domains.items():
            fill = ALT_FILL if alt else WHITE_FILL
            row_data = [
                record.student_name or "",
                record.class_sec or "",
                record.roll_no or "",
                month,
                domain,
                d.c1,
                d.c2,
                d.c3,
                d.c4,
                d.observational_anecdote or "",
                d.strengths or "",
                d.focus_next_month or "",
                sig(record.parent_sign_present),
                sig(record.teacher_sign_present),
                sig(record.principal_sign_present),
                pathlib.Path(record.source_pdf or "").name,
            ]
            ws.append(row_data)

            for col_idx, (col_name, value) in enumerate(zip(COLUMNS, row_data), 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = BORDER
                cell.font   = DATA_FONT

                if col_name in SCORE_COLS and isinstance(value, int):
                    color = SCORE_COLOR.get(value)
                    if color:
                        cell.fill = PatternFill("solid", fgColor=color)
                    cell.alignment = CENTER
                elif col_name in WIDE_COLS:
                    cell.fill      = fill
                    cell.alignment = LEFT
                else:
                    cell.fill      = fill
                    cell.alignment = CENTER

            ws.row_dimensions[current_row].height = 20
            current_row += 1
            alt = not alt

    return current_row


class ExcelWriter:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # remove default empty sheet
        self._sheets: dict[str, openpyxl.worksheet.worksheet.Worksheet] = {}
        self._next_row: dict[str, int] = {}

        for stage, sheet_name in SHEET_NAMES.items():
            ws = self.wb.create_sheet(title=sheet_name)
            _header_row(ws, COLUMNS)
            _col_widths(ws, COLUMNS)
            ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
            self._sheets[stage] = ws
            self._next_row[stage] = 2  # data starts row 2

    def add_record(self, record: LPCRecord):
        ws = self._sheets.get(record.stage)
        if not ws:
            raise ValueError(f"Unknown stage: {record.stage}")
        next_row = _write_record(ws, record, self._next_row[record.stage])
        self._next_row[record.stage] = next_row

    def save(self, output_path: Optional[str] = None, buf=None) -> str:
        if buf is not None:
            self.wb.save(buf)
            buf.seek(0)
            return ""
        if not output_path:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"output/LPC_Report_{ts}.xlsx"
        pathlib.Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)
        return output_path
